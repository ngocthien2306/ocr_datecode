"""
Các định dạng bảng của báo cáo: Excel, CSV, JSON.

Cùng một bộ bảng cho cả ba, dựng một lần ở `build_tables` rồi mỗi định dạng
chỉ lo cách ghi ra. Nhờ vậy ba file xuất cùng lúc không thể lệch số nhau.
"""

import csv
import io
import json
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent_app.core.i18n import t, treport_period
from agent_app.reports.html_report import _fmt_label


def build_tables(
    cfg: Dict[str, Any],
    summary: Dict[str, Any],
    timeseries: Dict[str, Any],
) -> Dict[str, Any]:
    """Các bảng phẳng của báo cáo, kèm phần thông tin kỳ."""
    picked = cfg.get("selectedRecipeIds") or []
    recipes = [r for r in summary["by_recipe"] if r["recipe_id"] in picked] if picked \
        else summary["by_recipe"]

    if picked:
        total = sum(r["total"] for r in recipes)
        passed = sum(r["pass"] for r in recipes)
        failed = sum(r["fail"] for r in recipes)
        rate = round(passed / total * 100, 2) if total else 0.0
    else:
        total, passed = summary["total"], summary["pass"]
        failed, rate = summary["fail"], summary["pass_rate"]

    gran = cfg["granularity"]
    by_time = [{
        "bucket": _fmt_label(pt["timestamp"], gran),
        "raw_bucket": pt["timestamp"],
        "total": pt["total"], "pass": pt["pass"], "fail": pt["fail"],
        "pass_rate": pt["pass_rate"],
    } for pt in timeseries["data"]]

    by_time_recipe = []
    for pt in timeseries["data"]:
        for r in pt["by_recipe"]:
            if picked and r["recipe_id"] not in picked:
                continue
            by_time_recipe.append({
                "bucket": _fmt_label(pt["timestamp"], gran),
                "recipe_name": r["recipe_name"],
                "total": r["total"], "pass": r["pass"], "fail": r["fail"],
                "pass_rate": r["pass_rate"],
            })

    return {
        "meta": {
            # `cfg["periodLabel"]` luôn là tiếng Anh (port từ
            # reportGenerator.ts), nên phải dịch để bản tiếng Việt không
            # ghi "Kỳ báo cáo: Today".
            "period_label": treport_period(cfg["periodLabel"]),
            "start": cfg["startDate"],
            "end": cfg["endDate"],
            "granularity": gran,
            "recipe_scope": (", ".join(r["recipe_name"] for r in recipes)
                             if picked else t("Tất cả recipe")),
            "generated_at": cfg["generatedAt"],
        },
        "overview": {"total": total, "pass": passed, "fail": failed, "pass_rate": rate},
        "by_recipe": recipes,
        # `by_camera` rỗng khi kỳ báo cáo dài — xem `data.build_summary`, phần
        # tách theo camera phải nạp trọn document nên bị tắt mặc định.
        "by_camera": summary.get("by_camera") or [],
        "by_time": by_time,
        "by_time_recipe": by_time_recipe,
    }


# ── JSON ─────────────────────────────────────────────────────────────────────

def to_json(tables: Dict[str, Any]) -> bytes:
    return json.dumps(tables, ensure_ascii=False, indent=2).encode("utf-8")


# ── CSV ──────────────────────────────────────────────────────────────────────

def to_csv(tables: Dict[str, Any]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    m = tables["meta"]
    o = tables["overview"]

    def block(title: str, header: List[str], rows: List[List[Any]]) -> None:
        w.writerow([title])
        w.writerow(header)
        w.writerows(rows)
        w.writerow([])

    block(t("# THÔNG TIN BÁO CÁO"), [t("Trường"), t("Giá trị")], [
        [t("Kỳ báo cáo"), m["period_label"]],
        [t("Từ"), m["start"]],
        [t("Đến"), m["end"]],
        [t("Mốc thời gian"), m["granularity"]],
        [t("Phạm vi recipe"), m["recipe_scope"]],
        [t("Xuất lúc"), m["generated_at"]],
    ])
    block(t("# TỔNG QUAN"), [t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
          [[o["total"], o["pass"], o["fail"], o["pass_rate"]]])
    block(t("# THEO THỜI GIAN"), [t("Mốc"), t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
          [[r["bucket"], r["total"], r["pass"], r["fail"], r["pass_rate"]] for r in tables["by_time"]])
    block(t("# THEO RECIPE"), ["Recipe", t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
          [[r["recipe_name"], r["total"], r["pass"], r["fail"], r["pass_rate"]] for r in tables["by_recipe"]])
    block(t("# THEO THỜI GIAN × RECIPE"), [t("Mốc"), "Recipe", t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
          [[r["bucket"], r["recipe_name"], r["total"], r["pass"], r["fail"], r["pass_rate"]]
           for r in tables["by_time_recipe"]])
    if tables["by_camera"]:
        block(t("# THEO CAMERA"), ["Camera", "Serial", t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
              [[c["camera_id"], c["serial_number"], c["total"], c["pass"], c["fail"], c["pass_rate"]]
               for c in tables["by_camera"]])

    # BOM để Excel nhận đúng UTF-8; thiếu nó thì tiếng Việt ra ký tự lạ.
    return "﻿".encode("utf-8") + buf.getvalue().encode("utf-8")


# ── Excel ────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1E2A3A")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=13)


def _sheet(wb: Workbook, name: str, header: List[str],
           rows: List[List[Any]], widths: List[int]) -> None:
    # Excel giới hạn tên sheet 31 ký tự và không cho các ký tự : \\ / ? * [ ]
    safe = name[:31]
    ws = wb.create_sheet(safe)
    ws.append(header)
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    # Đóng băng dòng tiêu đề để cuộn bảng dài vẫn thấy tên cột.
    ws.freeze_panes = "A2"


def to_xlsx(tables: Dict[str, Any]) -> bytes:
    m, o = tables["meta"], tables["overview"]
    wb = Workbook()
    ws = wb.active
    ws.title = t("Tổng quan")
    ws.append([t("BÁO CÁO SẢN XUẤT")])
    ws["A1"].font = _TITLE_FONT
    ws.append([])
    for k, v in ((t("Kỳ báo cáo"), m["period_label"]), (t("Từ"), m["start"]), (t("Đến"), m["end"]),
                 (t("Mốc thời gian"), m["granularity"]), (t("Phạm vi recipe"), m["recipe_scope"]),
                 (t("Xuất lúc"), m["generated_at"])):
        ws.append([k, v])
    ws.append([])
    ws.append([t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")])
    for cell in ws[ws.max_row]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
    ws.append([o["total"], o["pass"], o["fail"], o["pass_rate"]])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40

    _sheet(wb, t("Theo thời gian"), [t("Mốc"), t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
           [[r["bucket"], r["total"], r["pass"], r["fail"], r["pass_rate"]] for r in tables["by_time"]],
           [22, 12, 12, 12, 16])
    _sheet(wb, t("Theo recipe"), ["Recipe", t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
           [[r["recipe_name"], r["total"], r["pass"], r["fail"], r["pass_rate"]] for r in tables["by_recipe"]],
           [34, 12, 12, 12, 16])
    _sheet(wb, t("Thời gian x Recipe"), [t("Mốc"), "Recipe", t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
           [[r["bucket"], r["recipe_name"], r["total"], r["pass"], r["fail"], r["pass_rate"]]
            for r in tables["by_time_recipe"]],
           [22, 34, 12, 12, 12, 16])
    if tables["by_camera"]:
        _sheet(wb, t("Theo camera"), ["Camera", "Serial", t("Tổng"), "Pass", "Fail", t("Tỷ lệ pass (%)")],
               [[c["camera_id"], c["serial_number"], c["total"], c["pass"], c["fail"], c["pass_rate"]]
                for c in tables["by_camera"]],
               [22, 22, 12, 12, 12, 16])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
