"""
Dựng báo cáo SO SÁNH nhiều máy: HTML, PDF, Excel, CSV.

Khác với `generate_report` ở edge — cái đó là báo cáo của MỘT máy. Ở đây nội dung
chính là phần so sánh, nên không thể gom năm file của năm máy lại là xong.

Cấu trúc theo `docs/ui/10-report-templates.md` (Prompt 2):

  Tờ 1 — Tổng quan: banner phạm vi dữ liệu, KPI, nhịp theo ngày, bảng máy
  Tờ 2 — Vân tay kiểu lỗi (thứ DUY NHẤT so được giữa các máy chạy khác mặt
         hàng), cách đọc, phát hiện chính, rồi PHỤ LỤC THEO MÁY chia theo tuần

Hai tờ chứ không ba: phụ lục KHÔNG mở một tờ riêng bằng banner của nó. Làm thế
thì tờ phát hiện chính chỉ dùng nửa trên còn nửa dưới trắng — cho nội dung chảy
tiếp thì mỗi thẻ máy tự tìm chỗ, và hai thẻ vừa một trang A4. Với 5 máy, 7 ngày,
bản in ra 6 trang; chiều cao từng khối đo bằng cây hộp của WeasyPrint, không
phải ước lượng.

Tính toán nằm ở `aggregate.py`, biểu đồ ở `charts.py`. File này chỉ dựng trang.
Tách ba việc vì bản trước trộn cả ba trong một f-string, nên câu "tuần này tính
từ đâu tới đâu" chỉ trả lời được bằng cách đọc HTML.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from fleet_app.reports import aggregate as agg
from fleet_app.reports import charts

logger = logging.getLogger(__name__)

OUT_DIR = Path(__file__).resolve().parents[2] / "generated_reports"
OUT_DIR.mkdir(exist_ok=True)

FORMATS = ("html", "pdf", "excel", "csv")
GRANULARITIES = ("day", "week")

# Giữ tên cũ cho code ngoài đang import.
_PALETTE = charts.PALETTE


def _color_for(name: str) -> str:
    return charts.color_for(name)


# Định dạng số dùng chung với `aggregate.findings()` — một định nghĩa duy nhất,
# vì các câu phát hiện chính nằm ngay dưới bảng và phải viết số giống hệt.
_fmt = agg.fmt_num
_pct = agg.fmt_pct


def _esc(s: Any) -> str:
    return (str(s if s is not None else "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _img(b64: Optional[str], cls: str = "") -> str:
    return (f'<img class="{cls}" src="data:image/png;base64,{b64}">'
            if b64 else "")


_RATE_WORD = {"good": "trong ngưỡng", "watch": "cần theo dõi",
              "bad": "dưới ngưỡng", "none": "không có số"}


# ═══════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════

_CSS = """
@page {
  size: A4; margin: 13mm 12mm 15mm;
  @bottom-left  { content: "OCR Datecode · Fleet Service";
                  font-size: 8pt; color: #98a1b3; }
  @bottom-right { content: "Trang " counter(page) " / " counter(pages);
                  font-size: 8pt; color: #98a1b3; }
}
* { box-sizing: border-box; }
body { margin: 0; font: 10.5pt/1.5 "Helvetica Neue", Arial, sans-serif;
       color: #141821; -webkit-print-color-adjust: exact; print-color-adjust: exact; }

/* Bản xem trên màn hình được đóng khung như trang giấy, để người duyệt thấy
   đúng cái sẽ in ra. Bản in bỏ hết khung này (@media print ở dưới). */
@media screen {
  body { background: #eef0f4; padding: 22px 0; }
  .sheet { width: 210mm; min-height: 297mm; margin: 0 auto 20px; padding: 13mm 12mm;
           background: #fff; box-shadow: 0 2px 14px rgba(20,24,33,.14); }
}
@media print { .sheet { padding: 0; } .sheet + .sheet { break-before: page; } }

/* ── Đầu trang ─────────────────────────────────────────────────────────── */
.hd { background: #1e2a3a; color: #fff; padding: 13px 16px; border-radius: 6px;
      display: flex; align-items: flex-start; gap: 14px; }
.hd h1 { margin: 0; font-size: 16pt; letter-spacing: .2px; }
.hd .sub { font-size: 9pt; color: #b6c2d2; margin-top: 3px; }
.hd .mid { margin-left: auto; text-align: right; font-size: 9pt; color: #dbe3ec; }
.hd .mid b { display: block; font-size: 10.5pt; color: #fff; }
.hd .logo { width: 74px; height: 40px; border: 1px dashed #4a5a70; border-radius: 4px;
            display: flex; align-items: center; justify-content: center;
            font-size: 7pt; color: #8b9bb0; }

/* ── Banner phạm vi dữ liệu ────────────────────────────────────────────── */
/* Thiếu máy là chuyện phải đọc thấy TRƯỚC mọi con số, không phải một chú thích
   cuối trang: một bảng thiếu một máy trông vẫn hoàn toàn bình thường. */
.bn { margin: 11px 0 0; padding: 9px 13px; border-radius: 6px; font-size: 9.5pt;
      border: 1px solid; break-inside: avoid; }
.bn.ok  { background: #f0f7f2; border-color: #bcd9c5; color: #1e5233; }
.bn.bad { background: #fff5eb; border-color: #f0c69a; color: #8a3f10; }
.bn b { font-weight: 700; }
.bn ul { margin: 5px 0 0; padding-left: 17px; }

/* ── KPI ───────────────────────────────────────────────────────────────── */
.kpis { display: flex; gap: 8px; margin: 12px 0 0; }
.kpi { flex: 1; border: 1px solid #e3e7ee; border-radius: 7px; padding: 7px 10px; }
.kpi .k { font-size: 7.5pt; text-transform: uppercase; letter-spacing: .06em;
          color: #98a1b3; }
.kpi .v { font-size: 17pt; font-weight: 700; margin-top: 1px;
          font-variant-numeric: tabular-nums; }
.kpi .x { font-size: 8pt; color: #5f6878; }
.kpi.good { background: #f2f8f4; border-color: #bcd9c5; } .kpi.good .v { color: #1e5233; }
.kpi.watch{ background: #fdf8ec; border-color: #ebd6a3; } .kpi.watch .v { color: #7a5a12; }
.kpi.bad  { background: #fdf2f1; border-color: #eebcb7; } .kpi.bad  .v { color: #8f2f27; }
.kpi.none .v { color: #98a1b3; }

/* ── Mục, bảng ─────────────────────────────────────────────────────────── */
h2 { font-size: 12pt; margin: 14px 0 2px; padding-bottom: 4px;
     border-bottom: 2px solid #1e2a3a; break-after: avoid; }
/* Giữ tiêu đề, câu dẫn và biểu đồ đi cùng nhau. Không có mấy dòng này thì
   WeasyPrint tách tiêu đề ở cuối trang còn hình sang trang sau, và người đọc
   gặp một tiêu đề trống — đã xảy ra thật ở bản đầu. */
.lead { color: #5f6878; font-size: 9pt; margin: 6px 0 8px; break-after: avoid; }
.note { color: #5f6878; font-size: 8.5pt; font-style: italic; margin: 6px 0 0; }
img { width: 100%; display: block; margin: 6px 0 2px; break-inside: avoid; }

table { border-collapse: collapse; width: 100%; font-size: 9pt; margin-top: 6px; }
th, td { padding: 4px 7px; text-align: right; border-bottom: 1px solid #e9ecf2;
         font-variant-numeric: tabular-nums; }
th { background: #f5f7fa; font-size: 7.5pt; text-transform: uppercase;
     letter-spacing: .04em; color: #5f6878; font-weight: 700; white-space: nowrap; }
th:first-child, td:first-child { text-align: left; font-variant-numeric: normal; }
tr { break-inside: avoid; }
td.n { font-weight: 700; white-space: nowrap; }
td.n .sub { display: block; font-weight: 400; color: #98a1b3; font-size: 7.5pt; }
.muted { color: #5f6878; font-weight: 400; }
tr.tot td { font-weight: 700; background: #f5f7fa; border-top: 1.5px solid #1e2a3a; }
tr.off td { color: #98a1b3; background: #fafbfc; }
.hot { background: #fdf1ee; color: #8f2f27; font-weight: 700; }
.dot { display: inline-block; width: 7px; height: 7px; border-radius: 2px;
       margin-right: 5px; vertical-align: 1px; }
.st { font-size: 7.5pt; padding: 1px 5px; border-radius: 3px; white-space: nowrap; }
.st.good { background: #e6f2ea; color: #1e5233; }
.st.watch{ background: #fbf1d9; color: #7a5a12; }
.st.bad  { background: #fbe3e0; color: #8f2f27; }
.st.none { background: #eef0f4; color: #7a8394; }

/* ── Khối "đọc thế nào" và phát hiện ───────────────────────────────────── */
/* Hình vân tay + bảng nhiệt là MỘT khối. Bảng này là bảng so sánh giữa các máy;
   tách ra hai trang thì không còn so được, mà bảng lại chú thích bằng chú giải
   của hình ngay trên nó. Nếu không đủ chỗ, cả khối sang trang sau — thà thế còn
   hơn cắt đôi. */
.keep { break-inside: avoid; }

.box { border: 1px solid #e3e7ee; border-left: 3px solid #1e2a3a; border-radius: 5px;
       padding: 9px 12px; margin: 10px 0 0; font-size: 9pt; break-inside: avoid; }
.box h3 { margin: 0 0 4px; font-size: 9.5pt; }
.box p { margin: 3px 0; color: #3b4453; }
.finds { margin: 8px 0 0; }
.find { display: flex; gap: 9px; padding: 7px 0; border-top: 1px solid #eef0f4;
        break-inside: avoid; font-size: 9pt; }
.find .tag { flex: 0 0 auto; width: 4px; border-radius: 2px; }
.find.bad   .tag { background: #c2544d; }
.find.watch .tag { background: #ca8a04; }
.find.info  .tag { background: #6b8caf; }
.find.good  .tag { background: #2f6f4f; }
.find .do { display: block; color: #5f6878; font-size: 8.5pt; margin-top: 2px; }

/* ── Phụ lục theo máy ──────────────────────────────────────────────────── */
/* Thẻ không được tách trang: nửa thẻ ở cuối trang này, nửa ở đầu trang sau,
   người đọc mất luôn cái đang so. */
.mc { border: 1px solid #e3e7ee; border-radius: 7px; padding: 9px 12px;
      margin: 9px 0 0; break-inside: avoid; }
.mc.off { background: #fafbfc; border-style: dashed; }
.mc-hd { display: flex; align-items: baseline; gap: 9px; }
.mc-hd h3 { margin: 0; font-size: 12pt; }
.mc-hd .ln { font-size: 9pt; color: #5f6878; }
.mc-hd .dev { margin-left: auto; font-size: 8pt; color: #98a1b3; }
/* Ba số trên MỘT dòng, không phải ba ô có khung.
   Ba ô khung cao 28mm cho ba con số, và mỗi thẻ máy vì thế cao 118–135mm — quá
   cao để hai thẻ vừa một trang A4, nên phụ lục 5 máy ngốn 4 trang mà mỗi trang
   thừa 100–150mm trắng. Đo bằng cây hộp của WeasyPrint, không phải ước lượng.

   `>` chứ không phải khoảng trắng ở selector con: `.trio div` chọn luôn cả
   .k/.v/.x bên trong, nên mỗi ô từng hiện ra ba khung lồng nhau. */
.trio { display: flex; gap: 18px; margin: 7px 0 0; padding: 5px 0;
        border-top: 1px solid #eef0f4; border-bottom: 1px solid #eef0f4; }
.trio > div { display: flex; align-items: baseline; gap: 6px; }
.trio .k, .trio .x { white-space: nowrap; }
.trio .k { font-size: 7pt; text-transform: uppercase; letter-spacing: .05em;
           color: #98a1b3; }
.trio .v { font-size: 12pt; font-weight: 700; font-variant-numeric: tabular-nums;
           white-space: nowrap; }
.trio .x { font-size: 7.5pt; color: #5f6878; }
.why { font-size: 8.5pt; color: #5f6878; margin: 6px 0 0; line-height: 1.42; }
.foot { margin-top: 16px; padding-top: 7px; border-top: 1px solid #e3e7ee;
        color: #98a1b3; font-size: 8pt; }
"""


# ═══════════════════════════════════════════════════════════════════════════
# Các khối trang
# ═══════════════════════════════════════════════════════════════════════════

def _header(title: str, sub: str, data: Dict[str, Any]) -> str:
    return f"""<div class="hd">
  <div><h1>{_esc(title)}</h1><div class="sub">{_esc(sub)}</div></div>
  <div class="mid"><b>{_esc(data['period_label'])}</b>
    Lập lúc {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
  <div class="logo">logo</div>
</div>"""


def _banner(cov: Dict[str, Any]) -> str:
    total = cov.get("machines_total") or 0
    ok = cov.get("machines_ok") or 0
    if not total:
        # Không có máy nào thì `complete` vẫn True và banner cũ in "Đủ cả 0/0
        # máy" — một báo cáo trắng trơn trông như một báo cáo bình thường của
        # một nhà máy không sản xuất gì. Đây là trạng thái phải hét lên.
        return ('<div class="bn bad"><b>KHÔNG đọc được máy nào.</b> Báo cáo này '
                'rỗng vì tầng fleet không thấy máy nào trong tailnet, KHÔNG phải '
                'vì nhà máy không sản xuất. Kiểm tra Tailscale và tiến trình '
                'fleet service rồi lập lại.</div>')
    if cov.get("complete"):
        return (f'<div class="bn ok"><b>Đủ cả {ok}/{total} máy.</b> '
                f'Mọi con số dưới đây đọc từ chính các máy tại thời điểm lập báo cáo.</div>')
    items = "".join(
        f"<li><b>{_esc(m['machine'])}</b> — {_esc(m.get('reason') or 'không rõ lý do')}</li>"
        for m in (cov.get("machines_missing") or []) + (cov.get("machines_degraded") or []))
    return (f'<div class="bn bad"><b>KHÔNG đủ đội hình — chỉ có số của {ok}/{total} máy.</b>'
            f' Các máy thiếu vẫn có mặt trong báo cáo kèm lý do; con số tổng ở dưới'
            f' là tổng của {ok} máy, không phải của cả nhà máy.<ul>{items}</ul></div>')


def _kpis(tot: Dict[str, Any], views: List[Dict[str, Any]]) -> str:
    lvl = agg.rate_level(tot.get("pass_rate"))
    per_day = sum(v["per_day"] or 0 for v in views) or None
    return f"""<div class="kpis">
  <div class="kpi"><div class="k">Tổng sản lượng</div>
    <div class="v">{_fmt(tot.get('products'))}</div>
    <div class="x">{_fmt(per_day, 0)} sản phẩm/ngày toàn nhà máy</div></div>
  <div class="kpi"><div class="k">Đạt</div>
    <div class="v">{_fmt(tot.get('pass'))}</div><div class="x">&nbsp;</div></div>
  <div class="kpi"><div class="k">Không đạt</div>
    <div class="v">{_fmt(tot.get('fail'))}</div><div class="x">&nbsp;</div></div>
  <div class="kpi {lvl}"><div class="k">Tỉ lệ đạt chung</div>
    <div class="v">{_pct(tot.get('pass_rate'))}</div>
    <div class="x">{_RATE_WORD[lvl]} (ngưỡng {_pct(agg.RATE_GOOD, 0)})</div></div>
</div>"""


def _machine_table(views: List[Dict[str, Any]], tot: Dict[str, Any]) -> str:
    body = ""
    for v in views:
        dot = f'<span class="dot" style="background:{charts.color_for(v["machine"])}"></span>'
        name = (f'<td class="n">{dot}{_esc(v["machine"])}'
                f'<span class="sub">{_esc(v["line"] or "")}</span></td>')
        if not v["has_data"]:
            body += (f'<tr class="off">{name}<td colspan="5" class="muted">'
                     f'{_esc(v["error"] or "không có dữ liệu trong kỳ")}</td>'
                     f'<td><span class="st none">không có số</span></td></tr>')
            continue
        lvl = agg.rate_level(v["rate"])
        rec = v["recipes"][0]["name"] if v["recipes"] else None
        body += (f'<tr>{name}'
                 f'<td>{_fmt(v["total"])}</td>'
                 f'<td>{_fmt(v["per_day"], 1)}</td>'
                 f'<td>{_fmt(v["per_active_day"], 1)}</td>'
                 f'<td>{v["active_days"]}/{v["period_days"]}</td>'
                 f'<td>{_pct(v["rate"])}</td>'
                 f'<td class="muted">{_esc(rec or "—")}</td></tr>')
    per_day = sum(v["per_day"] or 0 for v in views) or None
    return f"""<table>
<thead><tr><th>Máy</th><th>Sản lượng</th><th>Mỗi ngày<br>(cả kỳ)</th>
  <th>Mỗi ngày<br>(có chạy)</th><th>Ngày<br>có chạy</th><th>Tỉ lệ đạt</th>
  <th>Recipe chính</th></tr></thead>
<tbody>{body}
<tr class="tot"><td>TỔNG</td><td>{_fmt(tot.get('products'))}</td>
  <td>{_fmt(per_day, 0)}</td><td>—</td><td>—</td>
  <td>{_pct(tot.get('pass_rate'))}</td><td class="muted">—</td></tr></tbody></table>
<p class="note">Không xếp hạng theo tỉ lệ đạt: các máy chạy recipe khác nhau, nên
tỉ lệ đạt phản ánh độ khó của mặt hàng chứ không phản ánh máy. Hai cột “mỗi ngày”
cố ý đặt cạnh nhau — máy chạy ít ngày trong kỳ thì cột đầu thấp hơn năng lực thật.</p>"""


def _fingerprint_block(fp: Dict[str, Any]) -> str:
    if not fp:
        return ('<p class="lead">Không có mẫu fail nào trong kỳ để dựng vân tay '
                'kiểu lỗi.</p>')
    labels = fp["labels"]
    # Nhãn NGẮN ở tiêu đề. Nhãn đầy đủ dài 30–38 ký tự, năm cột như thế rộng hơn
    # khổ A4 và bảng bị đẩy tràn khỏi lề — đo được ở bản đầu. Nhãn đầy đủ vẫn ở
    # chú giải biểu đồ ngay phía trên, nối với bảng bằng đúng ô màu đó.
    head = "".join(
        f'<th><span class="dot" style="background:{charts.cause_color(c, i)}"></span>'
        f'{_esc(charts.CAUSE_SHORT.get(c, labels.get(c, c)))}</th>'
        for i, c in enumerate(fp["causes"]))
    body = ""
    for r in fp["rows"]:
        cells = "".join(
            f'<td class="{"hot" if c["hot"] else ""}">{_pct(c["value"], 1)}</td>'
            for c in r["cells"])
        # Máy không có mẫu nào thì KHÔNG ghi "lấy mẫu" — "— · lấy mẫu" đọc như
        # là đã lấy mẫu mà không ra gì, còn thật ra là chưa lấy được mẫu nào.
        if not r["sample_products"]:
            note = ""
        else:
            note = " · " + ("phủ hết kỳ" if r["covers_all"] else "lấy mẫu")
        body += (f'<tr><td class="n">{_esc(r["machine"])}</td>{cells}'
                 f'<td class="muted">{_fmt(r["sample_products"])}{note}</td></tr>')
    med = "".join(f'<td class="muted">{_pct(fp["median"].get(c), 1)}</td>'
                  for c in fp["causes"])
    return f"""<div class="keep">{_img(charts.fingerprint(fp))}
<table><thead><tr><th>Máy</th>{head}<th>Mẫu</th></tr></thead>
<tbody>{body}
<tr class="tot"><td>Trung vị</td>{med}<td class="muted">—</td></tr></tbody></table>
<p class="note">Tên cột viết ngắn; tên đầy đủ ở chú giải biểu đồ phía trên, nối
bằng ô màu. Ô tô đỏ = cao hơn trung vị của cột từ 12 điểm trở lên. Trung vị chỉ
tính khi có từ 3 máy trở lên cùng có số cho cột đó — ít hơn thì “trung vị” là
chính con số duy nhất ấy, nên để “—”. Cột “Mẫu”
ghi cỡ mẫu của từng máy: “lấy mẫu” nghĩa là các tỉ trọng trên hàng đó là tỉ trọng
CỦA MẪU, không phải của toàn bộ số fail trong kỳ.</p></div>"""


_HOW_TO_READ = """<div class="box">
  <h3>Đọc vân tay thế nào</h3>
  <p><b>Detector không thấy vùng nào</b> cao → đi xem camera, trigger, ánh sáng.
     Máy không nhìn thấy chỗ cần đọc, nên chưa tới bước đọc.</p>
  <p><b>Ký tự dưới ngưỡng tin cậy</b> cao → máy THẤY vùng in nhưng không đủ tự tin
     về ký tự: nét in, mực, tiêu cự.</p>
  <p><b>OCR đọc sai chuỗi</b> cao → đọc rõ nhưng ra chuỗi khác chuỗi mong đợi:
     xem lại chuỗi khai trong recipe.</p>
  <p><b>Ảnh không khớp template</b> cao → template có thể đã cũ so với bao bì đang chạy.</p>
  <p>Hai máy cùng “tỉ lệ đạt thấp” có thể đang hỏng hai thứ hoàn toàn khác nhau —
     đó là lý do trang này tồn tại, và là thứ duy nhất so được giữa các máy chạy
     khác mặt hàng.</p>
</div>"""


def _findings_block(finds: List[Dict[str, Any]]) -> str:
    rows = "".join(
        f'<div class="find {f["kind"]}"><div class="tag"></div><div>{_esc(f["text"])}'
        f'<span class="do">→ {_esc(f["action"])}</span></div></div>'
        for f in finds)
    return f'<h2>Phát hiện chính</h2><div class="finds">{rows}</div>'


def _week_table(v: Dict[str, Any]) -> str:
    """Bảng theo tuần của một máy. Tuần chưa đủ 7 ngày được ghi rõ là chưa đủ."""
    weeks = [w for w in v["weeks"] if w["total"]]
    if not weeks:
        return '<p class="why">Không có ngày nào có sản lượng trong kỳ.</p>'
    body = ""
    for w in weeks:
        d = w["delta_points"]
        delta = ("—" if d is None else
                 f'{"▲" if d > 0 else "▼" if d < 0 else "▬"} {_fmt(abs(d), 2)} điểm')
        lvl = agg.rate_level(w["rate"])
        # Ghi NGÀY ĐỨNG MÁY, không ghi "chưa đủ tuần". Kỳ 7 ngày thì tuần nào
        # cũng bị kỳ cắt, nên dòng nào cũng mang chữ đó và nó hết mang tin;
        # còn "2 ngày không chạy" giữa khoảng đã có số thì là chuyện thật.
        idle = f' · {w["idle_days"]} ngày không chạy' if w["idle_days"] else ''
        body += (f'<tr><td class="n">{_esc(w["label"])}'
                 f'<span class="sub">{_esc(w["span"])}{idle}</span></td>'
                 f'<td>{_fmt(w["total"])}</td>'
                 f'<td>{_fmt(w["pass"])}</td>'
                 f'<td>{_fmt(w["fail"])}</td>'
                 f'<td>{w["days"]}</td>'
                 f'<td>{_fmt(w["per_active_day"], 0)}</td>'
                 f'<td>{_pct(w["rate"])}</td>'
                 f'<td class="muted">{delta}</td>'
                 f'<td><span class="st {lvl}">{_RATE_WORD[lvl]}</span></td></tr>')
    return f"""<table>
<thead><tr><th>Tuần</th><th>Sản lượng</th><th>Đạt</th><th>Không đạt</th>
  <th>Ngày<br>có chạy</th><th>Mỗi ngày<br>có chạy</th><th>Tỉ lệ đạt</th>
  <th>So tuần trước</th><th></th></tr></thead><tbody>{body}</tbody></table>"""


def _machine_card(v: Dict[str, Any]) -> str:
    dot = f'<span class="dot" style="background:{charts.color_for(v["machine"])}"></span>'
    head = (f'<div class="mc-hd"><h3>{dot}{_esc(v["machine"])}</h3>'
            f'<span class="ln">{_esc(v["line"] or "")}</span>'
            f'<span class="dev">{_esc(v["model"] or "thiết bị chưa khai")}</span></div>')

    if not v["has_data"]:
        # Máy thiếu số VẪN có thẻ. Bỏ nó ra là báo cáo trông đầy đủ trong khi
        # một dây chuyền của nhà máy không được nhắc tới ở đâu cả.
        return (f'<div class="mc off">{head}'
                f'<p class="why"><b>Không có dữ liệu trong kỳ.</b> '
                f'{_esc(v["error"] or "chưa rõ lý do")} — máy có thể vẫn đang sản '
                f'xuất; thiếu ở đây là thiếu ĐƯỜNG ĐỌC SỐ, không phải thiếu sản lượng.</p>'
                f'</div>')

    lvl = agg.rate_level(v["rate"])
    recs = ", ".join(f'{_esc(r["name"])} ({_fmt(r.get("products"))})'
                     for r in v["recipes"][:3]) or "—"
    top = v.get("top_cause") or {}
    why = []
    if top:
        why.append(f'Nguyên nhân lớn nhất trên mẫu: <b>{_esc(top.get("label"))}</b> '
                   f'({_fmt(top.get("share_of_causes_pct"), 0)}% của mẫu'
                   f'{" · " + _esc(v["sampling"]) if v.get("sampling") else ""}).')
    if v.get("worst_day") and v["worst_day"]["rate"] is not None:
        wd = v["worst_day"]
        why.append(f'Ngày thấp nhất: {wd["key"][8:10]}/{wd["key"][5:7]} — '
                   f'{_pct(wd["rate"])} trên {_fmt(wd["total"])} sản phẩm.')
    if v["active_days"] and v["active_days"] < v["period_days"]:
        why.append(f'Chỉ có sản lượng {v["active_days"]}/{v["period_days"]} ngày '
                   f'của kỳ.')

    return f"""<div class="mc">{head}
  <div class="trio">
    <div><div class="k">Sản lượng</div><div class="v">{_fmt(v['total'])}</div>
      <div class="x">{_fmt(v['per_active_day'], 0)}/ngày có chạy</div></div>
    <div><div class="k">Tỉ lệ đạt</div><div class="v">{_pct(v['rate'])}</div>
      <div class="x">{_RATE_WORD[lvl]}</div></div>
    <div><div class="k">Không đạt</div><div class="v">{_fmt(v['fail'])}</div>
      <div class="x">trên {_fmt(v['total'])} sản phẩm</div></div>
  </div>
  {_img(charts.machine_trend(v))}
  {_week_table(v)}
  <p class="why">Recipe trong kỳ: {recs}<br>{"<br>".join(why)}</p>
</div>"""


# ═══════════════════════════════════════════════════════════════════════════
# HTML
# ═══════════════════════════════════════════════════════════════════════════

def build_html(data: Dict[str, Any]) -> str:
    rows = data["machines"]
    days = data.get("period_days") or 7
    # Phải gọi TRƯỚC khi vẽ bất cứ hình nào: hình và ô màu trong bảng đọc cùng
    # một bảng gán, lệch nhau là mất đường nối giữa chúng.
    charts.configure([r["machine"] for r in rows])
    views = [agg.machine_view(r, days) for r in rows]
    tot = data["fleet_total"]
    fp = agg.fingerprint_view(data.get("failure_fingerprint") or {})
    cov = data["coverage"]
    finds = agg.findings(views, fp, cov)
    live = [v for v in views if v["has_data"]]

    sub = (f"{len(rows)} dây chuyền · {data['period_label']} · "
           f"nhà máy đóng gói gia vị")

    page1 = f"""<div class="sheet">
{_header('BÁO CÁO SO SÁNH DÂY CHUYỀN', sub, data)}
{_banner(cov)}
{_kpis(tot, views)}
<h2>Nhịp sản xuất theo ngày</h2>
<p class="lead">Cột chồng theo máy, nên một ngày sản lượng tụt là đọc được ngay
máy nào tụt — thứ mà bảng dưới không nói được.</p>
{_img(charts.fleet_daily(live))}
<h2>Sản lượng theo máy</h2>
{_machine_table(views, tot)}
</div>"""

    page2 = f"""<div class="sheet">
{_header('VÂN TAY KIỂU LỖI', sub, data)}
<h2>Tỉ trọng nguyên nhân trên mẫu fail</h2>
<p class="lead">Các ô của MỘT máy cộng lại bằng 100. Đây là cách so sánh có nghĩa
giữa các máy, vì những nguyên nhân này thuộc pipeline OCR chứ không thuộc mặt
hàng — khác với tỉ lệ đạt, thứ phản ánh độ khó của sản phẩm đang chạy.</p>
{_fingerprint_block(fp)}
{_HOW_TO_READ}
{_findings_block(finds)}"""

    cards = "".join(_machine_card(v) for v in views)
    page3 = f"""<h2>Sản lượng chuẩn hoá theo máy</h2>
<p class="lead">Mỗi máy hai cột: chia cho cả kỳ, và chia cho số ngày thực có
chạy. Máy chạy 2 trong 7 ngày mà chỉ đọc cột thứ nhất thì trông như máy yếu —
đọc cả hai mới ra “chạy nhanh nhưng chạy ít ngày”. Đây là câu trả lời cho “vì
sao máy này thấp”, nên nó đứng ngay trước phần chi tiết từng máy.</p>
{_img(charts.output_per_day(live))}
<h2>Phụ lục theo máy · chia theo tuần</h2>
<p class="lead">Mỗi máy một thẻ, chia theo tuần ISO (thứ Hai đầu tuần) — không
phải cửa sổ 7 ngày trượt, để hai báo cáo lập cách nhau một ngày vẫn so được với
nhau. Ngày đứng máy trong tuần được ghi ngay dưới tên tuần. Kỳ chỉ có một tuần
thì biểu đồ vẽ theo ngày, vì một cột đơn độc không so được với gì cả.</p>
{cards}
<div class="foot">Sinh bởi Fleet Service · dữ liệu đọc trực tiếp từ từng máy tại
thời điểm lập báo cáo · giá trị không đo được hiện “—”, không hiện 0.</div>"""

    return f"""<!doctype html><html lang="vi"><head><meta charset="utf-8">
<title>Báo cáo so sánh dây chuyền · {_esc(data['period_label'])}</title>
<style>{_CSS}</style></head><body>
{page1}
{page2}
{page3}
</div>
</body></html>"""


def build_pdf(html: str, path: Path) -> None:
    from weasyprint import HTML
    HTML(string=html).write_pdf(str(path))


# ═══════════════════════════════════════════════════════════════════════════
# Excel / CSV
# ═══════════════════════════════════════════════════════════════════════════

def build_excel(data: Dict[str, Any], path: Path) -> None:
    """
    Excel nhiều sheet: sản lượng · theo tuần · vân tay lỗi.

    Tách sheet thay vì nhồi một bảng: các bảng có ĐƠN VỊ khác nhau (số sản phẩm
    và tỉ trọng %), để chung một sheet là mời người đọc cộng nhầm cột.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    days = data.get("period_days") or 7
    views = [agg.machine_view(r, days) for r in data["machines"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sản lượng"
    ws.append(["Máy", "Dây chuyền", "Thiết bị", "Sản lượng", "Mỗi ngày (cả kỳ)",
               "Mỗi ngày (có chạy)", "Ngày có chạy", "Đạt", "Không đạt",
               "Tỉ lệ đạt (%)", "Recipe chính", "Ghi chú"])
    for v in views:
        ws.append([v["machine"], v["line"], v["model"], v["total"], v["per_day"],
                   v["per_active_day"], v["active_days"], v["pass"], v["fail"],
                   v["rate"], (v["recipes"][0]["name"] if v["recipes"] else None),
                   v["error"] or ""])
    t = data["fleet_total"]
    ws.append([])
    ws.append(["TỔNG", "", "", t.get("products"), "", "", "", t.get("pass"),
               t.get("fail"), t.get("pass_rate"), "", ""])

    w2 = wb.create_sheet("Theo tuần")
    w2.append(["Máy", "Tuần ISO", "Từ–đến", "Đủ 7 ngày", "Sản lượng", "Đạt",
               "Không đạt", "Ngày có chạy", "Mỗi ngày có chạy", "Tỉ lệ đạt (%)",
               "So tuần trước (điểm)"])
    for v in views:
        for w in v["weeks"]:
            if not w["total"]:
                continue
            w2.append([v["machine"], f"{w['year']}-W{w['week']:02d}", w["span"],
                       not w["partial"], w["total"], w["pass"], w["fail"],
                       w["days"], w["per_active_day"], w["rate"],
                       w["delta_points"]])
    if w2.max_row == 1:
        w2.append(["Không có tuần nào có sản lượng trong kỳ."])

    fp = agg.fingerprint_view(data.get("failure_fingerprint") or {})
    if fp:
        w3 = wb.create_sheet("Vân tay kiểu lỗi")
        labels = fp["labels"]
        w3.append(["Máy"] + [labels.get(c, c) for c in fp["causes"]]
                  + ["Cỡ mẫu", "Phủ hết kỳ"])
        for r in fp["rows"]:
            w3.append([r["machine"]] + [c["value"] for c in r["cells"]]
                      + [r["sample_products"], r["covers_all"]])
        w3.append(["Trung vị"] + [fp["median"].get(c) for c in fp["causes"]]
                  + [None, None])
        w3.append([])
        w3.append(["Các ô là TỈ TRỌNG giữa các nguyên nhân (một máy cộng lại = 100), "
                   "không phải % sản phẩm trượt từng bước."])

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E2A3A")
    for sheet in wb.worksheets:
        for c in sheet[1]:
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None),
                        default=8)
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 42)
        sheet.freeze_panes = "A2"

    wb.save(str(path))


def build_csv(data: Dict[str, Any], path: Path) -> None:
    """
    CSV một file, hai khối: tổng hợp theo máy rồi chi tiết theo tuần.

    Một file chứ không hai: người nhận mở bằng Excel để lọc, và hai file rời là
    hai lần phải nhớ file nào đi với file nào. Khối thứ hai có dòng tiêu đề
    riêng nên bộ lọc vẫn dùng được.
    """
    days = data.get("period_days") or 7
    views = [agg.machine_view(r, days) for r in data["machines"]]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["# Tổng hợp theo máy", data.get("period_label")])
        w.writerow(["Máy", "Dây chuyền", "Thiết bị", "Sản lượng",
                    "Mỗi ngày (cả kỳ)", "Mỗi ngày (có chạy)", "Ngày có chạy",
                    "Đạt", "Không đạt", "Tỉ lệ đạt (%)", "Recipe chính", "Ghi chú"])
        for v in views:
            w.writerow([v["machine"], v["line"], v["model"], v["total"],
                        v["per_day"], v["per_active_day"], v["active_days"],
                        v["pass"], v["fail"], v["rate"],
                        (v["recipes"][0]["name"] if v["recipes"] else None),
                        v["error"] or ""])
        t = data["fleet_total"]
        w.writerow(["TỔNG", "", "", t.get("products"), "", "", "", t.get("pass"),
                    t.get("fail"), t.get("pass_rate"), "", ""])
        w.writerow([])
        w.writerow(["# Chi tiết theo tuần"])
        w.writerow(["Máy", "Tuần ISO", "Từ–đến", "Đủ 7 ngày", "Sản lượng", "Đạt",
                    "Không đạt", "Ngày có chạy", "Mỗi ngày có chạy",
                    "Tỉ lệ đạt (%)", "So tuần trước (điểm)"])
        for v in views:
            for wk in v["weeks"]:
                if not wk["total"]:
                    continue
                w.writerow([v["machine"], f"{wk['year']}-W{wk['week']:02d}",
                            wk["span"], not wk["partial"], wk["total"],
                            wk["pass"], wk["fail"], wk["days"],
                            wk["per_active_day"], wk["rate"], wk["delta_points"]])


_EXT = {"html": "html", "pdf": "pdf", "excel": "xlsx", "csv": "csv"}


def render(data: Dict[str, Any], fmt: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    names = "-".join(r["machine"] for r in data["machines"])[:60]
    path = OUT_DIR / f"fleet_{names}_{stamp}.{_EXT[fmt]}"
    if fmt == "html":
        path.write_text(build_html(data), encoding="utf-8")
    elif fmt == "pdf":
        build_pdf(build_html(data), path)
    elif fmt == "excel":
        build_excel(data, path)
    else:
        build_csv(data, path)
    return path
